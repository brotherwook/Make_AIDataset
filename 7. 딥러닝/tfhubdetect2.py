import numpy as np
import cv2
import tensorflow_hub as hub
import os
import xml.etree.ElementTree as et
import sys
import openpyxl

# %%  모델 로드 및 xml 저장경로 설정 /캡쳐된 이미지 경로 설정
model = hub.load("https://tfhub.dev/tensorflow/faster_rcnn/resnet101_v1_800x1333/1")

excel_path = None

if len(sys.argv) == 1:
    video_path ='C:\MyWorkspace\Make_AIDataset\inputs\F20001;3_3sxxxx0;양재1동 23;양재환승센터;(Ch 01)_[20200928]080000-[20200928]080600(20200928_080000).avi'
    image_save_path = 'C:\MyWorkspace\Make_AIDataset\outputs\image'
    imagename = 'test'
elif len(sys.argv) == 2:
    excel_path = sys.argv[1]
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel['Sheet1']
elif len(sys.argv) == 4:
    video_path = sys.argv[1]
    image_save_path = sys.argv[2]
    imagename = sys.argv[3]
else :
    print("Usage: python person_detecting.py input_video_path(avi) output_image_dir output_image_name")
    print("원본 동영상 경로 및 저장할 저장할 배경이미지 경로(파일이름까지)(경로에 한글 포함되면 안됩니다.)")

#%%
def number(num):
 return '%03d' % num

#%% 들여쓰기 함수
def apply_indent(elem, level = 0):
    # tab = space * 2
    indent = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
        for elem in elem:
            apply_indent(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = indent

#%% 마우스 클릭을 위한 부분 (마스크 만들기)
clicked_points = []
clone = None

def MouseLeftClick(event, x, y, flags, param):

    image = clone.copy()

    # 왼쪽 마우스가 클릭되면 (x, y) 좌표를 저장한다.
    if event == cv2.EVENT_LBUTTONDOWN:
        clicked_points.append((x, y))

        # 원본 파일을 가져 와서 clicked_points에 있는 점들을 그린다.
        for point in clicked_points:
            cv2.circle(image, (point[0], point[1]), 2, (0, 255, 255), thickness=4)

        cv2.imshow("image", image)




# 새 윈도우 창을 만들고 그 윈도우 창에 MouseLeftClick 함수를 세팅해 줍니다.
cv2.namedWindow("image")
cv2.setMouseCallback("image", MouseLeftClick)


#%%
num = 0
cnt = 0

flag = 0
mask = None
roi = [[]]
temp = []

#%%
if excel_path == None:
    length = 2
else:
    length = sheet.max_row + 1

for i in range(2, length, 1):
    if excel_path is not None:
        video_path = sheet.cell(column=2, row=i).value
        image_save_path = sheet.cell(column=3, row=i).value
        imagename = video_path[21:42]
        print(video_path)
        print(image_save_path)
        print(imagename)

    #  영상불러오기
    cap = cv2.VideoCapture(video_path)
    if (not cap.isOpened()):
        print('Error opening video')

    total_frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
    height, width = (int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)), int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)))

    try:
        os.mkdir(image_save_path + "/" + imagename)
    except:
        print(imagename, "폴더가 이미 있음")
    try:
        os.mkdir(image_save_path + "/" + imagename + "/original")
    except:
        print(imagename + "/original 폴더가 이미있음")
    try:
        os.mkdir(image_save_path + "/" + imagename + "/roi")
    except:
        print(imagename + "/detect 폴더가 이미 있음")
    try:
        os.mkdir(image_save_path + "/" + imagename + "/detect")
    except:
        print(imagename + "/detect 폴더가 이미 있음")

    root = et.Element("annotations")

    while True:
        ret, img = cap.read()

        if flag == 0:
            while True:
                if mask is None:
                    # gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    mask = np.zeros_like(img)

                clone = cv2.resize(img, (int(width/2), int(height/2)))
                # clone = img.copy()
                cv2.imshow("image", clone)
                key = cv2.waitKey(0)

                # 클릭한거 저장 및 마스크 생성
                if key == ord('n'):
                    roi = [[]]
                    for points in clicked_points:
                        # print("(" + str(points[1]) + ", " + str(points[0]) + ')')
                        roi[0].append((points[0]*2, points[1]*2))
                    if len(roi[0]) > 0:
                        roi = np.array(roi)
                        cv2.fillPoly(mask, roi, (255, 255, 255))
                        clicked_points = []
                        temp.append(roi[0])

                # 클릭한거 취소
                if key == ord('b'):
                    if len(clicked_points) > 0:
                        clicked_points.pop()
                        image = clone.copy()
                        for point in clicked_points:
                            cv2.circle(image, (point[1], point[0]), 2, (0, 255, 255), thickness=-1)
                        cv2.imshow("image", image)

                if key == ord('r'):
                    cv2.destroyAllWindows()
                    clicked_points = []
                    flag = 1
                    print("roi 좌표:", roi)
                    break

        if flag == 1:
            if num % int(cap.get(cv2.CAP_PROP_FPS)) == 0:
                name = "/" + imagename + "_" + number(cnt) +".jpg"
                cv2.imwrite(image_save_path + "/" + imagename + "/original" + name, img)

                tag = et.SubElement(root, "image", {"id": str(num), "name": name, "width": str(width), "height": str(height)})

                if len(roi[0]) > 0:
                    # 마스크가 생성됬을 경우 마스크 처리한 부분만 조사
                    # cv2.imshow("mask", mask)
                    roi_img = cv2.bitwise_and(img, mask)
                    # roi 영역 빨간줄 긋기
                    for j in range(len(temp)):
                        for i, v in enumerate(roi[0]):
                            if i < len(roi[0]) - 1:
                                frame = cv2.line(img, tuple(temp[j][i]), tuple(temp[j][i + 1]), (0, 0, 255), 2)
                            else:
                                frame = cv2.line(img, tuple(temp[j][i]), tuple(temp[j][0]), (0, 0, 255), 2)

                    cv2.imwrite(image_save_path + "/" + imagename + "/roi" + name, frame)
                else:
                    roi_img = img

                cv2.imshow("roi_img", roi_img)

                input = cv2.cvtColor(roi_img, cv2.COLOR_BGR2RGB)
                input = input[np.newaxis, ...]
                output = model(input)

                threshold = 0.2

                for i, score in enumerate(output["detection_scores"][0]):
                    if score < threshold:
                        break
                    ymin, xmin, ymax, xmax = output["detection_boxes"][0][i]
                    (left, top) = (int(xmin * width), int(ymin * height))
                    (right, bottom) = (int(xmax * width), int(ymax * height))
                    if (right - left) * (bottom - top) > 10000:  # 너무 클 경우 무시
                        continue  # 영상마다 측정하여 수치변경
                    # if (right - left) * (bottom - top) < 1100:  # 너무 작은것도 무시
                    #     continue
                    if output["detection_classes"][0][i] == 1:  # <= 8: # 1~8 = 사람,자전거,승용차,오토바이,비행기,버스,전철,트럭
                        # 참고: https://github.com/amikelive/coco-labels/blob/master/coco-labels-paper.txt
                        img = cv2.rectangle(img, (left, top), (right, bottom), (255, 255, 255), 2)
                        # class_entity = "{:.2f}%".format(score * 100)
                        # cv2.putText(img_original, class_entity, (left, bottom + 10), cv2.FONT_HERSHEY_COMPLEX_SMALL, 0.4,
                        #             (0, 0, 0), 1)
                        et.SubElement(tag, "box", {"label": "사람", "occluded": "0", "source": "manual",
                                                   "xbr": str(right), "xtl": str(left),
                                                   "ybr": str(bottom), "ytl": str(top)})

                cv2.imwrite(image_save_path + "/" + imagename + '/detect' + name, img)
                cnt += 1
                img_tmp = cv2.resize(img, (int(width/2), int(height/2)))
                cv2.imshow("image", img_tmp)

            num += 1
            if cv2.waitKey(1) == 27:
                break

            if num % 4 == 0:
                loading = "\\"
            elif num % 4 == 1:
                loading = "|"
            elif num % 4 == 2:
                loading = "/"
            else:
                loading = "-"
            print("\r"+loading +"진행중 {:.2f}%".format((num / int(total_frames)) * 100), end='')


    cv2.destroyAllWindows()
    cap.release()

    apply_indent(root)

    xml = et.ElementTree(root)
    xml.write(image_save_path + "/" + imagename + "/" + imagename + ".xml")
    # flag = 0
    print("완료\n")
