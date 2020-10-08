import numpy as np
import cv2
import sys
from xml.etree.ElementTree import Element, SubElement, ElementTree
import os
import openpyxl

# cmd 창에서 실행시 python person_detecting.py input_video_path(avi) output_image_dir output_image_name 순으로 적어주면 됩니다.

#%% ========================================== 코드 설명 ====================================================
# cv2.createBackgroundSubtractorKNN() 함수를 사용해서 객체를 먼저 찾고
# 마우스클릭으로 못 찾은 객체나 잘못찾은 객체를 수정하는 코드입니다
# xml파일을 original 이미지 저장 경로에 같이 저장됩니다.
#
# 먼저 roi설정을 마우스 클릭으로 선택 후 'n'(대문자 인식안됩니다)을 누르면 mask가 저장이 됩니다
# 이후 r을 누르면 영상이 시작됩니다.
#
# 영상 시작 후
# 첫프레임은 일단 객체를 못찾습니다. 첫프레임만큼은 수동으로...
# 마우스로 클릭하면 설정한 크기의 박스가 생성됩니다. 원하는 곳에 박스를 만들고 's'를 누르면 해당 위치가 저장이 됩니다.
# 지우고 싶은 박스가 있는경우 그 박스의 가운데 점을 클릭 후 똑같이 's'를 눌러주면 박스가 사라집니다.
# 1,2,3번 키를 통해 사람,이륜차,자전거로 라벨명과 박스 크기를 변경할 수 있습니다.\
# 기준 박스크기 변경은 1,2,3번 키를 통해 바꾸고 오른쪽 마우스 드래그로 원하는 박스크기를 설정 후 스페이스바를 눌러주시면 변경이 됩니다.
# 해당 프레임의 수정이 끝났으면 'd'를 눌러 다음 프레임으로 넘어갈 수 있습니다.
# 영상은 1초에 1프레임씩 넘어갑니다.
#
# 재생 도중 멈추고싶으면 'esc'키를 눌러 종료하면 종료시점까지의 데이터(xml,사진)이 저장됩니다.
# ==========================================================================================================

#%% 사람영상 바뀔 때마다 바꿔줘야할 전역변수들
# 그냥 초기값 나중에 오른쪽마우스 드래그로 박스를 만들어서 값을 조정할 수 있습니다.
# 사람만 찾고 이륜차,자전거는 직접 해줘야합니다.

# 사람의 박스 크기
h_width = 20
h_height = 40

# 이륜차
motor_width = 60
motor_height = 60

# 자전거
bi_width = 110
bi_height = 130

# label 종류에 따른 변수 (초기값 사람에 맞춤)
label = '사람'
label_english = 'person'
label_width = h_width
label_height = h_height

half_width = int(label_width / 2)
half_height = int(label_height / 2)

excel_path = None

if len(sys.argv) == 1:
    video_path ='C:\MyWorkspace\Make_AIDataset\inputs\F20001;3_3sxxxx0;양재1동 23;양재환승센터;(Ch 01)_[20200928]080000-[20200928]080600(20200928_080000).avi'
    image_save_path = 'C:\MyWorkspace\Make_AIDataset\outputs\image'
    imagename = 'test'
elif len(sys.argv) == 2:
    excel_path = sys.argv[1]
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel['Sheet1']
elif len(sys.argv) == 4:
    video_path = sys.argv[1]
    image_save_path = sys.argv[2]
    imagename = sys.argv[3]
else :
    print("Usage: python person_detecting.py input_video_path(avi) output_image_dir output_image_name")
    print("원본 동영상 경로 및 저장할 저장할 배경이미지 경로(파일이름까지)(경로에 한글 포함되면 안됩니다.)")




#%% object class
class AllObject:
    def __init__(self, blobs):
        self.all_object = []
        for i,v in enumerate(blobs):
            person = Person(v)
            self.all_object.append(person)

    def update(self, object):
        self.all_object.append(object)

class Person:
    def __init__(self, position):
        self.label = '사람'
        self.center = position
        self.width = h_width
        self.height = h_height


class Motorcycle:
    def __init__(self, position):
        self.label = '이륜차'
        self.center = position
        self.width = motor_width
        self.height = motor_height


class Bicycle:
    def __init__(self, position):
        self.label = '자전거'
        self.center = position
        self.width = bi_width
        self.height = bi_height


#%% 마우스 클릭을 위한 부분 (마스크 만들기)
clicked_points = []
clone = None
click = False
x1,y1 = -1,-1
Rbox = None

def MouseLeftClick(event, x, y, flags, param):
    global flag
    global click
    global x1, y1
    global Rbox

    image = clone.copy()

    # 왼쪽 마우스가 클릭되면 (x, y) 좌표를 저장한다.
    if event == cv2.EVENT_LBUTTONDOWN:
        clicked_points.append((x, y))

        # 원본 파일을 가져 와서 clicked_points에 있는 점들을 그린다.
        for point in clicked_points:
            cv2.circle(image, (point[0], point[1]), 2, (0, 255, 255), thickness=4)
            cv2.rectangle(image, (point[0] - half_width, point[1] - half_height),
                          (point[0] + half_width, point[1] + half_height), (255, 255, 255), 2)
        cv2.imshow("image", image)

        if flag == 1:
            compare((x,y), blobs, label)

    elif event == cv2.EVENT_RBUTTONDOWN:
        # 마우스를 누른 상태
        click = True
        x1, y1 = x,y
        # print("사각형의 왼쪽위 설정 : (" + str(x1) + ", " + str(y1) + ")")

    elif event == cv2.EVENT_MOUSEMOVE:
        # 마우스 이동
        # 마우스를 누른 상태 일경우
        if click == True:
            cv2.rectangle(image,(x1,y1),(x,y),(255,0,0), 2)
            # print("(" + str(x1) + ", " + str(y1) + "), (" + str(x) + ", " + str(y) + ")")
            cv2.imshow("image", image)

    elif event == cv2.EVENT_RBUTTONUP:
        # 마우스를 때면 상태 변경
        click = False
        cv2.rectangle(image,(x1,y1),(x,y),(255,0,0), 2)
        cv2.imshow("image", image)
        Rbox = [int(x-x1), int(y-y1)]


# 새 윈도우 창을 만들고 그 윈도우 창에 MouseLeftClick 함수를 세팅해 줍니다.
cv2.namedWindow("image")
cv2.setMouseCallback("image", MouseLeftClick)

#%% 영상 전처리
def imagepreprocessing(img):
    width_min = h_width - 40
    width_max = h_width + 40
    height_min = h_height - 20
    height_max = sys.maxsize

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    fgmask = fgbg.apply(gray)

    # 그림자제거
    ret, thresh = cv2.threshold(fgmask, 127, 255, cv2.THRESH_BINARY)

    # 노이즈 제거
    medianblur = cv2.medianBlur(thresh, 5)

    # 팽창연산
    k = cv2.getStructuringElement(cv2.MORPH_CROSS, (2, 2))
    dil = cv2.dilate(medianblur, k)

    # 윤곽선찾기
    contours, hierachy = cv2.findContours(dil, mode, method)

    # 윤곽선 안을 흰색으로 채우기
    for i, contour in enumerate(contours):
        dil = cv2.fillPoly(dil, contour, 255)

    # 다시한번 팽창 연산
    k = cv2.getStructuringElement(cv2.MORPH_RECT, (4, 4))
    dil = cv2.dilate(dil, k)

    kernel = np.ones((5, 5), np.uint8)
    dil = cv2.morphologyEx(dil.astype(np.uint8), cv2.MORPH_CLOSE, kernel)

    # 다시 윤곽선 찾기
    contours, hierachy = cv2.findContours(dil, mode, method)

    # 중심점 구하기
    for i, contour in enumerate(contours):
        dil = cv2.fillPoly(dil, contour, 255)
        area = cv2.contourArea(contour)
        x, y, width, height = cv2.boundingRect(contour)

        if width_min <= width <= width_max and height_min <= height <= height_max:
            center = (int(x + width / 2), int(y + height / 2))
            blobs.append(center)

    # cv2.imshow("aa", dils)

    return dil

#%% xml 만드는 부분
root = None

def initXML():
    global root
    root = Element('annotations')

    SubElement(root, 'version').text = '1.1'

    meta = SubElement(root, 'meta')
    task = SubElement(meta, 'task')
    SubElement(task, 'id').text = '12937'
    SubElement(task, 'name').text = 'image_F18006_2_202009140900'
    SubElement(task, 'size').text = '365'
    SubElement(task, 'mode').text = 'annotation'
    SubElement(task, 'overlap').text = '0'
    SubElement(task, 'bugtracker').text
    SubElement(task, 'created').text = '2020-09-15 07:01:49.569020+01:00'
    SubElement(task, 'updated').text = '2020-09-16 08:37:29.588763+01:00'
    SubElement(task, 'start_frame').text = '0'
    SubElement(task, 'stop_frame').text = '364'
    SubElement(task, 'frame_filter').text
    SubElement(task, 'z_order').text = 'False'

    sub = SubElement(task, 'labels')
    ssub = SubElement(sub, 'label')
    SubElement(ssub, 'name').text = '사람'
    SubElement(ssub, 'color').text = '#1dc799'
    SubElement(ssub, 'attributes').text

    sub2 = SubElement(task, 'segments')
    ssub2 = SubElement(sub2, 'segment')
    SubElement(ssub2, 'id').text = '28'
    SubElement(ssub2, 'start').text = '0'
    SubElement(ssub2, 'stop').text = '364'
    SubElement(ssub2, 'url').text = 'http://222.107.208.162:8088/?id=28'


    sub3 = SubElement(task, 'owner')
    SubElement(sub3, 'username').text = 'brotherwook'
    SubElement(sub3, 'email').text = 'jahanda@naver.com'

    SubElement(task, 'assignee')

    SubElement(meta, 'dumped').text = '2020-09-16 08:37:30.049858+01:00'

def makeXML(id, name, all_object):
    global root

    image = root.find('image[@id="' + str(id) + '"]')
    if image is None:
        image = SubElement(root, 'image')
        image.attrib["id"] = str(id)
        image.attrib["name"] = name
        image.attrib["width"] = str(width)
        image.attrib["height"] = str(height)

    for i, v in enumerate(all_object):
        box = SubElement(image, 'box')
        box.attrib["label"] = v.label
        box.attrib["occluded"] = '0'
        box.attrib["source"] = 'manual'

        xtl = v.center[0] - v.width / 2
        ytl = v.center[1] - v.height / 2
        xbr = v.center[0] + v.width / 2
        ybr = v.center[1] + v.height / 2

        if xtl*2 < 0:
            xtl = 0
        if ytl*2 < 0:
            ytl = 0
        if xbr*2 > width:
            xbr = width
        if ybr*2 > height:
            ybr = height

        box.attrib["xtl"] = str(xtl*2)
        box.attrib["ytl"] = str(ytl*2)
        box.attrib["xbr"] = str(xbr*2)
        box.attrib["ybr"] = str(ybr*2)

# 들여쓰기 함수
def apply_indent(elem, level = 0):
    # tab = space * 2
    indent = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
        for elem in elem:
            apply_indent(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = indent

def number(num):
 return '%03d' % num

#%%
def compare(click, blobs, label):
    global create
    global removed
    global margin
    for j, center in enumerate(blobs):
        if (click[0] - margin < center[0] < click[0] + margin) and (click[1] - margin < center[1] < click[1] + margin):
            removed.append(center)
            removed.append(click)

    if click not in removed:
        create.append(click)

    for i, v in enumerate(create):
        blobs.append(v)
        if label.__eq__('사람'):
            obj = Person(v)
            everything.update(obj)
        elif label.__eq__('이륜차'):
            obj = Motorcycle(v)
            everything.update(obj)
        elif label.__eq__('자전거'):
            obj = Bicycle(v)
            everything.update(obj)

    # clicks[:] = clicks[1:]
    create = []

def removeBOX():
    global removed
    global blobs
    global everything

    # removed에 포함되어 있지 않은 것들만 저장
    blobs[:] = [v for j, v in enumerate(blobs)
                if v not in removed]

    everything.all_object[:] = [v for j, v in enumerate(everything.all_object)
                                if v.center not in removed]
    
    # 초기화
    removed = []

#%% 그 외 전역변수
fgbg = cv2.createBackgroundSubtractorKNN()


# 현재 프레임 번호 변수
t = 0

mode = cv2.RETR_EXTERNAL
method = cv2.CHAIN_APPROX_SIMPLE


blobs = []
margin = 10
person_counter = None
removed = []
create = []
id = 0

flag = 0
mask = None
roi = [[]]
temp = []

color = np.random.randint(0, 255, (200, 3))


# 동영상 저장용 (안씀)
# fourcc = cv2.VideoWriter_fourcc(*'mp4v')  # 디지털 미디어 포맷 코드 생성 , 인코딩 방식 설
# out = cv2.VideoWriter('C:/MyWorkspace/Make_AIDataset/outputs/video/output.avi', fourcc, cap.get(cv2.CAP_PROP_FPS), (width, height))

everything = None
cnt = 0
#%%
if excel_path == None:
    length = 2
else:
    length = sheet.max_row + 1

for i in range(2, length, 1):
    if excel_path is not None:
        video_path = sheet.cell(column=2, row=i).value
        image_save_path = sheet.cell(column=3, row=i).value
        imagename = video_path[21:42]
        print(video_path)
        print(image_save_path)
        print(imagename)

    #  영상불러오기
    cap = cv2.VideoCapture(video_path)
    if (not cap.isOpened()):
        print('Error opening video')

    total_frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
    height, width = (int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)), int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)))

    try:
        os.mkdir(image_save_path + "/" + imagename)
    except:
        print(imagename, "폴더가 이미 있음")
    try:
        os.mkdir(image_save_path + "/" + imagename + "/original")
    except:
        print(imagename + "/original 폴더가 이미있음")
    try:
        os.mkdir(image_save_path + "/" + imagename + "/roi")
    except:
        print(imagename + "/detect 폴더가 이미 있음")
    try:
        os.mkdir(image_save_path + "/" + imagename + "/detect")
    except:
        print(imagename + "/detect 폴더가 이미 있음")

    root = Element('annotations')
    initXML()
    while True:
        ret, frameo = cap.read()
        if not ret:
            break
        a = int(width/2)
        b = int(height/2)
        frame = cv2.resize(frameo, (a,b))
        # 마스크 설정 부분
        # 마우스로 보기를 원하는 부분 클릭하고 n누르면 해당부분만 확인
        # 원하는 부분 클릭후  다음 프레임에서 또 다시 클릭하면 모두 확인가능
        # r 누르면 영상 재생
        if flag == 0:
            while True:
                clone = frame.copy()
                cv2.imshow("image", frame)
                key = cv2.waitKey(0)

                if mask is None:
                    # gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    mask = np.zeros_like(frame)

                # 클릭한거 저장 및 마스크 생성
                if key == ord('n'):
                    roi = [[]]
                    for points in clicked_points:
                        # print("(" + str(points[1]) + ", " + str(points[0]) + ')')
                        roi[0].append((points[0], points[1]))
                    if len(roi[0]) > 0:
                        roi = np.array(roi)
                        cv2.fillPoly(mask, roi, (255, 255, 255))
                        clicked_points = []
                        temp.append(roi[0])

                # 클릭한거 취소
                if key == ord('b'):
                    if len(clicked_points) > 0:
                        clicked_points.pop()
                        image = clone.copy()
                        for point in clicked_points:
                            cv2.circle(image, (point[1], point[0]), 2, (0, 255, 255), thickness=-1)
                        cv2.imshow("image", image)

                if key == 32:
                    if Rbox is not None:
                        if label.__eq__('사람'):
                            h_width = int(Rbox[0])
                            h_height = int(Rbox[1])
                            label_width = h_width
                            label_height = h_height
                        elif label.__eq__('이륜차'):
                            motor_width = int(Rbox[0])
                            motor_height = int(Rbox[1])
                            label_width = motor_width
                            label_height = motor_height
                        elif label.__eq__('자전거'):
                            bi_width = int(Rbox[0])
                            bi_height = int(Rbox[1])
                            label_width = bi_width
                            label_height = bi_height
                    Rbox = None

                if key == ord('r'):
                    cv2.destroyAllWindows()
                    clicked_points = []
                    flag = 1
                    print("roi 좌표:", roi)
                    break


        if flag == 1:
            blobs = []

            # 새 윈도우 창을 만들고 그 윈도우 창에 MouseLeftClick 함수를 세팅해 줍니다.
            cv2.namedWindow("image")
            cv2.setMouseCallback("image", MouseLeftClick)
            name = "/" + imagename + "_" + number(cnt) +".jpg"

            if t % int(cap.get(cv2.CAP_PROP_FPS)) == 0:
                cv2.imwrite(image_save_path + "/" + imagename + "/original" + name, frameo)

            # ==============객체 찾는 부분 ===============
            # 객체 찾는 부분만 수정해서 사용하면 됩니다
            # 객체찾는 함수에서 center값만 blobs에 append해주면 됩니다
            if len(roi[0]) > 0:
                # 마스크가 생성됬을 경우 마스크 처리한 부분만 조사
                # cv2.imshow("mask", mask)
                roi_img = cv2.bitwise_and(frame, mask)

                # 객체 찾는 함수
                dil = imagepreprocessing(roi_img)

                # roi 영역 빨간줄 긋기
                for j in range(len(temp)):
                    for i,v in enumerate(roi[0]):
                        if i < len(roi[0])-1:
                            frame = cv2.line(frame,tuple(temp[j][i]),tuple(temp[j][i+1]), (0, 0, 255), 2)
                        else:
                            frame = cv2.line(frame, tuple(temp[j][i]), tuple(temp[j][0]), (0, 0, 255), 2)
            else:
                # 마스크를 따로 안만들면 영상 전체 조사
                # 객체 찾는 함수
                dil = imagepreprocessing(frame)

            #==================================================

            # 1초당 한프레임만 저장
            if t % int(cap.get(cv2.CAP_PROP_FPS)) == 0:
                everything = AllObject(blobs)
                # roi 빨간색으로 경계 쳐진 이미지 저장
                img_tmp = cv2.resize(frame, (width,height))
                cv2.imwrite(image_save_path + "/" + imagename + "/roi" + name, img_tmp)
                clone = frame.copy()
                cv2.putText(clone, label_english, (100, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2, cv2.LINE_AA)
                half_width = int(label_width / 2)
                half_height = int(label_height / 2)

                for i,obj in enumerate(everything.all_object):
                    cv2.rectangle(clone,(obj.center[0] - int(obj.width/2), obj.center[1] - int(obj.height/2)),
                                  (obj.center[0]+int(obj.width/2), obj.center[1]+int(obj.height/2)), (255, 255, 255), 2)
                    cv2.circle(clone, (obj.center[0], obj.center[1]), 2, (255, 0, 0), thickness=2)


                cv2.imshow("image", clone)
                cv2.imshow("dil", dil)

                key = cv2.waitKey(1)

                if key == ord("a") or t == 0:
                    while True:
                        clone = frame.copy()
                        cv2.putText(clone, label_english, (100, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2,
                                    cv2.LINE_AA)
                        half_width = int(label_width / 2)
                        half_height = int(label_height / 2)

                        for i, obj in enumerate(everything.all_object):
                            cv2.rectangle(clone, (obj.center[0] - int(obj.width / 2), obj.center[1] - int(obj.height / 2)),
                                          (obj.center[0] + int(obj.width / 2), obj.center[1] + int(obj.height / 2)),
                                          (255, 255, 255), 2)
                            cv2.circle(clone, (obj.center[0], obj.center[1]), 2, (255, 0, 0), thickness=2)

                        cv2.imshow("image", clone)

                        k = cv2.waitKey(0)
                        # 프로그램 종료
                        if k == 27:
                            flag = 2
                            break

                        # 다음 프레임으로 넘어가기
                        if k == ord("d"):
                            break

                        if k == ord('s'):
                            removeBOX()
                            clicked_points = []

                        if k == 32:
                            if Rbox is not None:
                                if label.__eq__('사람'):
                                    h_width = int(Rbox[0])
                                    h_height = int(Rbox[1])
                                    label_width = h_width
                                    label_height = h_height
                                elif label.__eq__('이륜차'):
                                    motor_width = int(Rbox[0])
                                    motor_height = int(Rbox[1])
                                    label_width = motor_width
                                    label_height = motor_height
                                elif label.__eq__('자전거'):
                                    bi_width = int(Rbox[0])
                                    bi_height = int(Rbox[1])
                                    label_width = bi_width
                                    label_height = bi_height
                            Rbox = None

                        # 사람 1번
                        if k == 49:
                            label = '사람'
                            label_english = 'person'
                            label_width = h_width
                            label_height = h_height

                        # 이륜차 2번
                        if k == 50:
                            label = '이륜차'
                            label_english = 'motorcycle'
                            label_width = motor_width
                            label_height = motor_height

                        # 자전거 3번
                        if k == 51:
                            label = '자전거'
                            label_english = 'bicycle'
                            label_width = bi_width
                            label_height = bi_height

                elif key == 27:
                    flag = 2
                    break

                makeXML(id, name, everything.all_object)
                id += 1
                clicked_points = []
                blobs = []
                # detecting한 이미지 저장
                img_tmp = cv2.resize(clone, (width,height))
                cv2.imwrite(image_save_path + "/" + imagename + '/detect' + name, img_tmp)
                cnt += 1

        else:
            break

        t += 1


    cap.release()
    # out.release()
    cv2.destroyAllWindows()
    # flag = 0

    #%% xml파일 생성

    # 들여쓰기
    apply_indent(root)

    # xml 파일로 보내기
    tree = ElementTree(root)
    tree.write(image_save_path + "/" + imagename + "/" + imagename + ".xml")



